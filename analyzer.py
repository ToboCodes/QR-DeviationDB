import mysql.connector
import pandas as pd
import sys
import datetime
import os
from openpyxl.styles import Alignment, Border, Side

def fetch_data(start_date):
    db_config = {
        'host': 'localhost',
        'user': 'root',
        'password': os.environ.get('DB_PASSWORD'),
        'database': 'desviaciones'
    }

    # Convert date to UNIX time in ms
    unix_timestamp = int(datetime.datetime.strptime(start_date, '%Y-%m-%d').timestamp() * 1000)

    query = """
        SELECT 
            o.offense_id, d.name AS domain_name, o.category_count, o.device_count, o.event_count, 
            o.local_destination_count, o.magnitude, o.username_count, o.query_time, o.source_count,
            o.remote_destination_count
        FROM 
            Offense o
        JOIN 
            Domains d ON o.domain_id = d.domain_id
        WHERE 
            o.offense_id IN (
                SELECT DISTINCT offense_id
                FROM Offense
                WHERE last_updated_time >= %s
            )
        ORDER BY 
            o.offense_id, o.query_time;
    """

    with mysql.connector.connect(**db_config) as connection:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(query, (unix_timestamp,))
        data = cursor.fetchall()

    return data

def analyze_data(data):
    results = []

    for i in range(0, len(data) - 4, 5):
        for j in range(0, 4):
            current_data = data[i + j]
            next_data = data[i + j + 1]

            if current_data["offense_id"] != next_data["offense_id"]:
                break

            # Extracting common fields
            common_fields = {
                'offense_id': current_data['offense_id'],
                'domain_name': current_data['domain_name'],
                'query_time': next_data['query_time']
            }

            columns_to_check = ["category_count", "device_count", "local_destination_count", "magnitude", 
                                "username_count", "source_count", "remote_destination_count"]

            for column in columns_to_check:
                delta_value = next_data[column] - current_data[column]
                if delta_value > 0:
                    results_entry = {**common_fields, 'column': column, 'delta_value': delta_value}
                    results.append(results_entry)

            event_count_delta = next_data['event_count'] - current_data['event_count']
            if event_count_delta >= 50 and next_data['query_time'] != 4:
                results_entry = {**common_fields, 'column': 'event_count', 'delta_value': event_count_delta}
                results.append(results_entry)

    return results

def generate_summary(results):
    summary = {}
    
    magnitude_map = {
        'category_count': lambda x: x['delta_value'] * 2,
        'device_count': lambda x: 5 if x['delta_value'] > 2 else 0,
        'event_count': lambda x: 3 if x['delta_value'] >= 50 and x['query_time'] != 4 else 0,
        'local_destination_count': 3,
        'magnitude': 2,
        'username_count': 1,
        'source_count': 3,
        'remote_destination_count': 2
    }

    for row in results:
        key = (row['domain_name'], row['offense_id'])
        if key not in summary:
            summary[key] = {'evolution': set(), 'magnitude': 0}
        
        summary[key]['evolution'].add(row['query_time'])

        if callable(magnitude_map.get(row['column'])):
            summary[key]['magnitude'] += magnitude_map[row['column']](row)
        else:
            summary[key]['magnitude'] += magnitude_map.get(row['column'], 0)

    # List of dictionaries for easier document writing
    summary_list = [{'domain_name': key[0], 'offense_id': key[1], 'evolution': len(val['evolution']), 'magnitude': val['magnitude']} for key, val in summary.items()]
    
    return sorted(summary_list, key=lambda x: x['domain_name'])


def auto_adjust_columns_width(sheet):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

def write_to_excel(results, summary, filename):
    results_df = pd.DataFrame(results).drop(columns=['domain_name'])  # Drop domain_name column
    summary_df = pd.DataFrame(summary)
    
    # Column renaming
    renamed_columns_data = {
        'offense_id': 'Ofensa',
        'column': 'Atributo',
        'delta_value': 'Delta',
        'query_time': 'Muestra'
    }
    
    renamed_columns_summary = {
        'domain_name': 'Cliente',
        'offense_id': 'Ofensa',
        'evolution': 'Evolución',
        'magnitude': 'Magnitud'
    }

    results_df.rename(columns=renamed_columns_data, inplace=True)
    summary_df.rename(columns=renamed_columns_summary, inplace=True)

    with pd.ExcelWriter(f'Reporte Mutaciones {filename}.xlsx', engine='openpyxl') as writer:
        # Reordering sheets - 'Resumen' first, 'Data' next
        summary_df.sort_values(by='Magnitud', ascending=False, inplace=True)
        summary_df.to_excel(writer, sheet_name='Resumen', index=False)

        results_df.to_excel(writer, sheet_name='Data', index=False)
        
        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            last_col_letter = sheet.cell(row=1, column=sheet.max_column).column_letter
            sheet.auto_filter.ref = f"A1:{last_col_letter}1"
            
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = Border(left=Side(style='thin'), 
                                         right=Side(style='thin'), 
                                         top=Side(style='thin'), 
                                         bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='left')
            auto_adjust_columns_width(sheet)

def main():
    if len(sys.argv) != 2:
        print("Usage: script.py YYYY-MM-DD")
        sys.exit(1)

    start_date = sys.argv[1]
    data = fetch_data(start_date)
    results = analyze_data(data)
    summary = generate_summary(results)
    
    # Write Excel with two sheets
    write_to_excel(results, summary, start_date)

if __name__ == '__main__':
    main()